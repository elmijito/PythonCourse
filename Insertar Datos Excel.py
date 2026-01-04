import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# --- Crear o cargar libro de Excel ---
archivo_excel = "datos_usuarios.xlsx"

def cargar_o_crear_excel():
    """Carga el archivo si existe, o crea uno nuevo con encabezados."""
    if os.path.exists(archivo_excel):
        wb = load_workbook(archivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre", "Edad", "Email", "Telefono", "Dirección"])
        wb.save(archivo_excel)  # Guarda el archivo nuevo con encabezados
    return wb, ws

# Cargar o crear el Excel al iniciar
wb, ws = cargar_o_crear_excel()

# --- Función para guardar los datos del formulario ---
def guardar_datos():
    # Obtener los valores del formulario
    nombre = entry_nombre.get().strip()
    edad = entry_edad.get().strip()
    email = entry_Email.get().strip()
    telefono = entry_telefono.get().strip()
    direccion = entry_direccion.get().strip()

    # Validar campos vacíos
    if not (nombre and edad and email and telefono and direccion):
        messagebox.showwarning("Campos vacíos", "Por favor completa todos los campos antes de guardar.")
        return

    # Recargar el archivo antes de escribir (por si fue modificado externamente)
    wb, ws = cargar_o_crear_excel()

    # Agregar nueva fila
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(archivo_excel)  # Guarda los cambios en el archivo

    # Limpiar campos del formulario
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_Email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)

    # Mostrar mensaje de éxito
    messagebox.showinfo("Éxito", "Los datos fueron guardados correctamente en el archivo Excel.")

# --- Interfaz gráfica ---
root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg='#4B6587')

label_style = {"bg": '#4B6587', "fg": "white"}
entry_style = {"bg": '#D3D3D3', "fg": "black"}

# Campos del formulario
tk.Label(root, text="Nombre:", **label_style).grid(row=0, column=0, padx=10, pady=5, sticky="e")
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Edad:", **label_style).grid(row=1, column=0, padx=10, pady=5, sticky="e")
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Email:", **label_style).grid(row=2, column=0, padx=10, pady=5, sticky="e")
entry_Email = tk.Entry(root, **entry_style)
entry_Email.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Teléfono:", **label_style).grid(row=3, column=0, padx=10, pady=5, sticky="e")
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

tk.Label(root, text="Dirección:", **label_style).grid(row=4, column=0, padx=10, pady=5, sticky="e")
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

# Botón Guardar
boton_guardar = tk.Button(root, text="Guardar", bg='#6D8299', fg='white', command=guardar_datos)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
